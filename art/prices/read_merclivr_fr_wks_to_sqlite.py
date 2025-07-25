#!/usr/bin/env python3
"""

    wksfilepath='example.ods',
    sqlitepath='output.db',
    tablename='ods_data',

"""
import argparse
import sqlite3
from openpyxl import load_workbook
from odf import opendocument
from odf.table import Table, TableRow, TableCell
import os
# from typing import Dict, List, Union
parser = argparse.ArgumentParser(description="Tranform from worksheet to Sqlite")
parser.add_argument("--wksfilepath", type=str,
                    help="Path to worksheet file")
parser.add_argument("--sqlitepath", type=str, default=None,
                    help="Path to Sqlite file")
parser.add_argument("--tablename", type=str, default="tablename",
                    help="SQL tablename")
args = parser.parse_args()


class SpreadsheetToSQLite:
  """
  A class to convert spreadsheet tables (Excel or OpenOffice) to SQLite databases.

  Features:
  - Reads Excel (.xlsx) and OpenOffice (.ods) files
  - Introspects headers as column names
  - Creates appropriate SQLite tables
  - Handles data type inference
  - Preserves data relationships

  Usage:
  converter = SpreadsheetToSQLite()
  converter.convert('input.xlsx', 'output.db', table_name='my_table')
  """

  def __init__(self):
    self.header = []
    self.rows = []
    self.column_types = {}

  def _read_excel(self, file_path: str, sheet_name: str = None) -> None:
    """Read data from Excel file (.xlsx)"""
    wb = load_workbook(filename=file_path, read_only=True)

    # Use the first sheet if not specified
    sheet = wb[sheet_name] if sheet_name else wb.active

    # Read header (first row)
    self.header = [str(cell.value).strip() for cell in next(sheet.iter_rows())]

    # Read remaining rows
    self.rows = []
    for row in sheet.iter_rows(min_row=2):
      row_data = []
      for cell in row:
        # Convert cell value to the appropriate Python type
        value = cell.value
        if isinstance(value, str):
          value = value.strip()
        row_data.append(value)
      self.rows.append(row_data)

    self._infer_column_types()

  def _read_ods(self, file_path: str, table_name: str = None) -> None:
    """Read data from OpenOffice file (.ods)"""
    doc = opendocument.load(file_path)

    # Find the right table (use first if not specified)
    tables = doc.getElementsByType(Table)
    if not tables:
      raise ValueError("No tables found in ODS file")

    table = tables[0] if table_name is None else None
    if table_name is not None:
      for t in tables:
        if t.getAttribute('name') == table_name:
          table = t
          break
      if table is None:
        raise ValueError(f"Table '{table_name}' not found in ODS file")

    # Read header (first row)
    header_row = table.getElementsByType(TableRow)[0]
    self.header = []
    for cell in header_row.getElementsByType(TableCell):
      # Get text content of cell
      text_content = []
      for t in cell.childNodes:
        if t.nodeType == 1:  # Element node
          text_content.append(t.data)
      cell_value = ''.join(text_content).strip()
      self.header.append(cell_value)

    # Read remaining rows
    self.rows = []
    for row in table.getElementsByType(TableRow)[1:]:
      row_data = []
      for cell in row.getElementsByType(TableCell):
        # Get text content of cell
        text_content = []
        for t in cell.childNodes:
          if t.nodeType == 1:  # Element node
            text_content.append(t.data)
        cell_value = ''.join(text_content).strip()

        # Try to convert to the appropriate type
        try:
          cell_value = int(cell_value)
        except ValueError:
          try:
            cell_value = float(cell_value)
          except ValueError:
            pass
        row_data.append(cell_value)
      self.rows.append(row_data)

    self._infer_column_types()

  def _infer_column_types(self) -> None:
    """Infer SQLite column types from the data"""
    if not self.rows:
      return

    # Initialize with first row's types
    self.column_types = {}
    for i, value in enumerate(self.rows[0]):
      col_name = self.header[i]
      if isinstance(value, int):
        self.column_types[col_name] = 'INTEGER'
      elif isinstance(value, float):
        self.column_types[col_name] = 'REAL'
      else:
        self.column_types[col_name] = 'TEXT'

    # Refine with further rows
    for row in self.rows[1:]:
      for i, value in enumerate(row):
        if i >= len(self.header):
          continue  # Skip if row has more columns than header

        col_name = self.header[i]
        current_type = self.column_types[col_name]

        if current_type == 'INTEGER':
          if not isinstance(value, int):
            if isinstance(value, float):
              self.column_types[col_name] = 'REAL'
            else:
              self.column_types[col_name] = 'TEXT'
        elif current_type == 'REAL':
          if not (isinstance(value, (int, float))):
            self.column_types[col_name] = 'TEXT'

  def read_spreadsheet(self, file_path: str, sheet_or_table_name: str = None) -> None:
    """
    Read spreadsheet data from a file.

    Args:
        file_path: Path to the spreadsheet file
        sheet_or_table_name: Name of a sheet (Excel) or table (OpenOffice) to read
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.xlsx':
      self._read_excel(file_path, sheet_or_table_name)
    elif ext == '.ods':
      self._read_ods(file_path, sheet_or_table_name)
    else:
      raise ValueError(f"Unsupported file format: {ext}")

  def write_sqlite(self, db_path: str, table_name: str, if_exists: str = 'replace') -> None:
    """
    Write the loaded spreadsheet data to an SQLite database.

    Args:
        db_path: Path to SQLite database file
        table_name: Name of table to create
        if_exists: What to do if table exists ('replace', 'append', 'fail')
    """
    if not self.header:
      raise ValueError("No data loaded. Call read_spreadsheet() first.")

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Create table SQL
    columns_sql = []
    for col in self.header:
      col_type = self.column_types.get(col, 'TEXT')
      columns_sql.append(f'"{col}" {col_type}')

    create_table_sql = f"CREATE TABLE IF NOT EXISTS \"{table_name}\" ({', '.join(columns_sql)})"

    if if_exists == 'replace':
      cursor.execute(f"DROP TABLE IF EXISTS \"{table_name}\"")
    elif if_exists == 'fail' and self._table_exists(cursor, table_name):
      conn.close()
      raise ValueError(f"Table '{table_name}' already exists and if_exists='fail'")

    cursor.execute(create_table_sql)

    # Insert data
    placeholders = ', '.join(['?'] * len(self.header))
    insert_sql = f"INSERT INTO \"{table_name}\" VALUES ({placeholders})"

    for row in self.rows:
      # Ensure row has the same number of columns as header
      if len(row) < len(self.header):
        row = list(row) + [None] * (len(self.header) - len(row))
      elif len(row) > len(self.header):
        row = row[:len(self.header)]
      cursor.execute(insert_sql, row)
    conn.commit()
    conn.close()

  @staticmethod
  def _table_exists(cursor: sqlite3.Cursor, table_name: str) -> bool:
    """Check if a table exists in the SQLite database"""
    cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    return cursor.fetchone() is not None

  def convert(
      self,
      wksfilepath: str,
      sqlitepath: str,
      tablename: str = 'tablename',
      sheetname: str = None,
      if_exists: str = None
    ) -> None:
    """
    Convenience method to read and write in one step.

    Args:
        wksfilepath: Path to input spreadsheet file
        sqlitepath: Path to output SQLite database
        tablename: Name of table to create in SQLite
        sheetname: Name of sheet/table in spreadsheet
        if_exists: What to do if table exists ('replace', 'append', 'fail')
    """
    if if_exists is None:
      if_exists = 'replace'
    self.read_spreadsheet(wksfilepath, sheetname)
    self.write_sqlite(sqlitepath, tablename, if_exists)


def get_args():
  """
  comm
    --wksfilepath "<path>"
    --sqlitepath "<path>"

  """
  wksfilepath = args.wksfilepath
  sqlitepath = args.sqlitepath
  return wksfilepath, sqlitepath


def usage_example():
  """
  """
  wksfilepath, sqlitepath = get_args()
  if sqlitepath is None:
    fopath, _ = os.path.split(wksfilepath)
    sqlitepath = os.path.join(fopath, 'sqlite.db')
  converter = SpreadsheetToSQLite()
  # Convert Excel to SQLite
  converter.convert(
    wksfilepath=wksfilepath,
    sqlitepath=sqlitepath
  )


def adhoctest():
  pass


def process():
  usage_example()


if __name__ == '__main__':
  """
  adhoctest()
  """
  process()
